import openpyxl

path="C:\\Users\Rohit\OneDrive\Desktop\data.xlsx"

workbook=openpyxl.load_workbook(path)

#sheet=workbook.get_sheet_by_name("sheet1")   #when we have multiple sheet
sheet=workbook.active
rows=sheet.max_row
columns=sheet.max_column

print("Total rows:",rows)
print("Total columns:",columns)

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(row=r,column=c).value,end="      ")
    print()